import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io


# Colour palette
CLR = {
    'green_hdr':  '1A7A4A',
    'green_row':  'D6F5E3',
    'red_hdr':    'C0392B',
    'red_row':    'FDECEA',
    'orange_hdr': 'D35400',
    'orange_row': 'FEF0E7',
    'blue_hdr':   '1565C0',
    'blue_row':   'E3F2FD',
    'grey_hdr':   '37474F',
    'grey_row':   'ECEFF1',
    'white':      'FFFFFF',
    'yellow':     'FFF9C4',
}

def _fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def _font(bold=False, color='000000', size=11):
    return Font(bold=bold, color=color, size=size)

def _border():
    thin = Side(style='thin', color='CCCCCC')
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _write_sheet(ws, df, hdr_color, row_color, col_subset=None, rename=None):
    """Write a dataframe to a worksheet with formatting."""
    rename = rename or {}
    if col_subset:
        df = df[col_subset].copy()
    df = df.rename(columns=rename)

    # Header row
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = _fill(hdr_color)
        cell.font = _font(bold=True, color='FFFFFF', size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = _border()

    ws.row_dimensions[1].height = 30

    # Data rows
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = _fill(row_color if row_idx % 2 == 0 else CLR['white'])
            cell.font = _font(size=9)
            cell.alignment = Alignment(vertical='center', wrap_text=False)
            cell.border = _border()

    # Auto width
    for col_idx, col_name in enumerate(df.columns, 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(col_name)),
            *[len(str(v)) for v in df.iloc[:, col_idx-1].fillna('')]
        ) if len(df) > 0 else len(str(col_name))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 35)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions


def _summary_sheet(ws, summary):
    rows = [
        ('GST ITC RECONCILIATION REPORT', '', ''),
        ('', '', ''),
        ('METRIC', 'COUNT', 'ITC AMOUNT (₹)'),
        ('Total invoices in GSTR-2B',        summary['total_2b'],            '—'),
        ('Total invoices in Books',           summary['total_books'],         '—'),
        ('✅ Matched (both sides)',           summary['matched'],             '—'),
        ('⚠️  Format Mismatch (same invoice, key diff)', summary['format_mismatch'], '—'),
        ('❌ In 2B — NOT in Books',           summary['not_in_books_count'],  f"{summary['not_in_books_itc']:,.0f}"),
        ('❌ In Books — NOT in 2B',           summary['not_in_2b_count'],     f"{summary['not_in_2b_itc']:,.0f}"),
        ('', '', ''),
        ('TOTAL ITC AT RISK (2B not in Books)',   '—', f"{summary['not_in_books_itc']:,.0f}"),
        ('TOTAL ITC AT RISK (Books not in 2B)',   '—', f"{summary['not_in_2b_itc']:,.0f}"),
        ('TOTAL NET ITC EXPOSURE',               '—', f"{summary['not_in_books_itc'] + summary['not_in_2b_itc']:,.0f}"),
    ]

    for r_idx, (a, b, c) in enumerate(rows, 1):
        for c_idx, val in enumerate([a, b, c], 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = _border()

            # Title row
            if r_idx == 1:
                cell.font = _font(bold=True, color='FFFFFF', size=14)
                cell.fill = _fill(CLR['grey_hdr'])
                cell.alignment = Alignment(horizontal='center', vertical='center')
            # Header row
            elif r_idx == 3:
                cell.font = _font(bold=True, color='FFFFFF', size=10)
                cell.fill = _fill(CLR['grey_hdr'])
                cell.alignment = Alignment(horizontal='center')
            # Matched
            elif r_idx == 6:
                cell.fill = _fill(CLR['green_row'])
                cell.font = _font(bold=(c_idx == 1), size=10)
            # Format mismatch
            elif r_idx == 7:
                cell.fill = _fill(CLR['yellow'])
                cell.font = _font(bold=(c_idx == 1), size=10)
            # Risk rows
            elif r_idx in (8, 9):
                cell.fill = _fill(CLR['red_row'])
                cell.font = _font(bold=(c_idx == 1), size=10)
            # Total rows
            elif r_idx >= 11:
                cell.fill = _fill(CLR['orange_row'])
                cell.font = _font(bold=True, size=10)
            else:
                cell.font = _font(size=10)

    ws.merge_cells('A1:C1')
    ws.row_dimensions[1].height = 35
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 22


def export_excel(results, output_path):
    writer = pd.ExcelWriter(output_path, engine='openpyxl')

    # Write dummy sheet first so we can use openpyxl after
    pd.DataFrame().to_excel(writer, sheet_name='Summary', index=False)
    writer.close()

    wb = load_workbook(output_path)

    # ── Summary ───────────────────────────────────────────────────────────────
    ws_sum = wb['Summary']
    _summary_sheet(ws_sum, results['summary'])

    # ── 2B Not In Books ───────────────────────────────────────────────────────
    ws_nib = wb.create_sheet('2B Not In Books ❌')
    df = results['not_in_books'].copy()
    show_cols = [c for c in df.columns if not c.startswith('_')]
    _write_sheet(ws_nib, df, CLR['red_hdr'], CLR['red_row'], col_subset=show_cols)

    # ── Books Not In 2B ───────────────────────────────────────────────────────
    ws_ni2b = wb.create_sheet('Books Not In 2B ❌')
    df2 = results['not_in_2b'].copy()
    show_cols2 = [c for c in df2.columns if not c.startswith('_')]
    _write_sheet(ws_ni2b, df2, CLR['red_hdr'], CLR['red_row'], col_subset=show_cols2)

    # ── Format Mismatch ───────────────────────────────────────────────────────
    ws_fm = wb.create_sheet('Format Mismatch ⚠️')
    df3 = results['fmt_mismatch'].copy()
    show_cols3 = [c for c in df3.columns if not c.startswith('_')]
    _write_sheet(ws_fm, df3, CLR['orange_hdr'], CLR['orange_row'], col_subset=show_cols3)

    # ── Matched ───────────────────────────────────────────────────────────────
    ws_m = wb.create_sheet('Matched ✅')
    df4 = results['matched'].copy()
    show_cols4 = [c for c in df4.columns if not c.startswith('_')]
    _write_sheet(ws_m, df4, CLR['green_hdr'], CLR['green_row'], col_subset=show_cols4)

    wb.save(output_path)
    return output_path
